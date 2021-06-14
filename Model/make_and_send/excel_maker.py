import openpyxl as excel
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime as dt

from Model.data_manager import DataManager


def adjust_column_width_from_col(ws, min_row, min_col, max_col):
    column_widths = []

    for i, col in \
            enumerate(
                ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row)
            ):

        for cell in col:
            value = cell.value
            if value is not None:

                if isinstance(value, str) is False:
                    value = str(value)

                try:
                    column_widths[i] = max(column_widths[i], len(value))
                except IndexError:
                    column_widths.append(len(value))

    for i, width in enumerate(column_widths):
        col_name = get_column_letter(min_col + i)
        value = column_widths[i] + 20
        ws.column_dimensions[col_name].width = value


def get_right_size(wi, he):
    ratio = wi / he
    if ratio > 0:
        he, wi = 100, 100 * ratio
    else:
        wi, he = 100, 100 * ratio
    return int(wi), int(he)


def export_excel(temp_data, temp_data_not, temp_name):
    write_data = excel.Workbook()
    worksheet = write_data.active
    worksheet.title = "Sheet1"
    title_font = Font(size=27, bold=True)
    normal_font = Font(size=20)

    alignment = Alignment(horizontal="center", vertical="center")

    worksheet.append(("   Employee ID   ", "     Employee Name     ", "    Unit Of Work    ",
                      "    Arrived Time    ", "    Using backup     ", "          Image          "))
    a = worksheet['A1:F1']
    for temp in a[0]:
        temp.font = title_font

    adjust_column_width_from_col(worksheet, 1, 1, worksheet.max_column)
    inDexCell = 1

    for index, tempData in enumerate(temp_data):
        # sub a space to make from "2021-06-13 02:30:22" to "2021-06-13_02h30m22"
        date_obj = dt.strptime(tempData[3], '%H:%M:%S')
        format_pic_name = dt.strftime(date_obj, '%Hh%Mm%S')

        if tempData[4] == 1:
            user_picture = "View/Backup/" + format_pic_name + ".jpg"
        else:
            user_picture = "View/Detected/" + format_pic_name + ".jpg"

        imgExcel = excel.drawing.image.Image(user_picture)
        imgExcel.height, imgExcel.width = get_right_size(wi=imgExcel.height, he=imgExcel.width)
        imgExcel.anchor = f'F{inDexCell + 1}'
        worksheet.merge_cells(f'F{inDexCell + 1}:F{inDexCell + 10}')
        worksheet.add_image(imgExcel)

        worksheet.append((
            tempData[0], tempData[1], tempData[2], tempData[3], tempData[4]
        ))

        a = worksheet[f'A{inDexCell + 2}:F{inDexCell + 2}']
        inDexCell += 10

        for temp in a[0]:
            temp.font = normal_font
            temp.alignment = alignment
        if tempData[4] == 1:
            ft = Font(color="4e42f5", size=20)
            for temp in a[0]:
                temp.font = ft
                temp.alignment = alignment
    for index, tempData in enumerate(temp_data_not):
        a = worksheet[f'A{inDexCell + 1}:C{inDexCell + 1}']
        inDexCell += 1

        ft = Font(color="00FF0000", size=20)
        for temp in a[0]:
            temp.font = ft
            temp.alignment = alignment
        worksheet.append((
            tempData[0], tempData[1], tempData[2]
        ))
    a = worksheet[f'A{inDexCell + 1}:C{inDexCell + 1}']
    ft = Font(color="00FF0000", size=20)
    for temp in a[0]:
        temp.font = ft
        temp.alignment = alignment

    write_data.save(filename=temp_name)


def export_excel_by_time(temp_data, temp_name):
    write_data = excel.Workbook()
    worksheet = write_data.active
    worksheet.title = "Sheet1"
    title_font = Font(size=25, bold=True)
    normal_font = Font(size=20)

    alignment = Alignment(horizontal="center", vertical="center")

    worksheet.append(("  Stt  ", " Employee ID ", "  Employee Name  ", " Unit Of Work ",
                      " All Session To Go ", "  Not Attendance  ", " Attendance Complete "))
    a = worksheet['A1:G1']
    for temp in a[0]:
        temp.font = title_font

    adjust_column_width_from_col(worksheet, 1, 1, worksheet.max_column)
    inDexCell = 1

    for index, tempData in enumerate(temp_data):

        worksheet.append((
            index + 1, tempData[0], tempData[1], tempData[2], tempData[3], tempData[4], tempData[5]
        ))

        inDexCell += 1
        a = worksheet[f'A{inDexCell}:G{inDexCell}']
        for temp in a[0]:
            temp.font = normal_font
            temp.alignment = alignment
        ft = Font(color="00FF0000", size=20)
        a[0][5].font = ft

    write_data.save(filename=temp_name)


def create_excel(ID_Recorder):
    # secret_infor = read_token_and_pass()
    # print(secret_infor,"<- word later on")

    with DataManager('Model/data/database/database.db') as db:
        data_export = db.get_list_to_export(ID_RECORDER=ID_Recorder)
        data_export2 = db.get_employee_did_not_go_to_work(ID_RECORDER=ID_Recorder)
    # sub a space to make from "2021-06-13 02:30:22" to "2021-06-13_02h30m22"
    date_obj = dt.strptime(ID_Recorder, '%Y-%m-%d %H:%M:%S')
    format_excel_name = dt.strftime(date_obj, '%Y-%m-%d_%Hh%Mm%S')

    excel_name_file = "View\\Summary\\" + format_excel_name + ".xlsx"
    export_excel(temp_data=data_export, temp_data_not=data_export2, temp_name=excel_name_file)
    return format_excel_name


def create_excel_by_time(begin_time, end_time):
    with DataManager('Model/data/database/database.db') as db:
        export_data = db.get_employee_to_export_by_time(begin_time, end_time)

    # # sub a space to make from "2021-06-13 02:30:22" to "2021-06-13_02h30m22"
    filename = begin_time + "to" + end_time
    excel_name = "View\\Summary\\" + filename + ".xlsx"
    export_excel_by_time(export_data, excel_name)
    return filename
